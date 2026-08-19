# ------------------
# version 1
# -------------------

from urllib.parse import urlencode
import requests
from pathlib import Path
from io import BytesIO
from copy import copy
from urllib.parse import urlencode, urljoin
from concurrent.futures import ThreadPoolExecutor, as_completed

import re
import requests
import demjson3
import pandas as pd

from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font


headers = {
    "User-Agent": (
        "Mozilla/5.0 (Linux; Android 15; Pixel 9) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/149.0.0.0 Mobile Safari/53.36"
    )
}

# ------------------------------------
# Create session for workers
# ------------------------------------

def create_session():
    """
    Create a requests session with retry and connection settings.

    Returns:
        requests.Session: Configured HTTP session.
    """
    session = requests.Session()
    retry=Retry(
        total=5,
        backoff_factor=1,
        status_forcelist=[
            429,502,503,504
        ],
        allowed_methods=["GET"]
    )

    adapter=HTTPAdapter(
        max_retries=retry
    )

    session.mount( "https://",adapter)
    session.headers.update(headers)
    return session


# -------------------------------------
# function for worker
# -------------------------------------

def get_datasource(item,sheet_id):
    """
    Fetch a financial statement from Codal and extract its datasource.

    Args:
        item (dict): Report data containing title and URLs.
        sheet_id (int): Financial statement sheet ID.

    Returns:
        dict: Result containing status, metadata, URLs, and datasource.
    """


    session = create_session()

    statement_url = item["url"] + f"&sheetid={sheet_id}"

    try:

            response = session.get(statement_url,timeout=30)
            response.raise_for_status()

    except requests.exceptions.RequestException as e:

        return {
            "status": "failed",
            "title": item["title"],
            "url": statement_url,
            "error": str(e)
        }

    html = response.text

    match = re.search(
        r"var\s+datasource\s*=\s*(\{.*?\})\s*;",
        html,
        re.DOTALL
    )

    if not match:

        return {
            "status": "no_datasource",
            "title": item["title"],
            "url": statement_url
        }



    try:
        datasource = demjson3.decode(match.group(1))

    except Exception as e:

            return {
                "status": "decode_error",
                "title": item["title"],
                "url": statement_url,
                "error": str(e)
            }

    return {
        "status": "success",
        "title": item["title"],
        "url": statement_url,
        "pdf": item["pdf_url"],
        "excel": item["excel_url"],
        "attachment": item.get("attachment"),
        "datasource": datasource
    }




def get_financial_report(
        symbol,
        statement_name="صورت سود و زیان",
        length=12,
        from_date=None,
        to_date=None,
        mains=True,
        consolidatable=True,
        NotConsolidatable=True,
        NotAudited=False,
        Audited=True,
    ):
    """
        Fetch financial reports from Codal and extract statement data.

        Args:
            symbol (str): Company symbol.
            statement_name (str): Financial statement name.
            length (int): Number of reports to fetch.
            from_date (str, optional): Start date for filtering reports.
            to_date (str, optional): End date for filtering reports.
            mains (bool): Include main reports.
            consolidatable (bool): Include consolidatable reports.
            NotConsolidatable (bool): Include non-consolidatable reports.
            NotAudited (bool): Include unaudited reports.
            Audited (bool): Include audited reports.

        Returns:
            list[dict]: Successfully extracted financial reports.
    """

    # -------------------------------
    # Create API with your filtering
    # -------------------------------

    params = {
        "Audited": str(Audited).lower(),
        "AuditorRef": -1,
        "Category": 1,
        "Childs": "false",
        "CompanyState": -1,
        "CompanyType": -1,
        "Consolidatable": str(consolidatable).lower(),
        "IsNotAudited": "false",
        "Length": length,
        "LetterType": 6,
        "Mains": str(mains).lower(),
        "NotAudited": str(NotAudited).lower(),
        "NotConsolidatable":str(NotConsolidatable).lower(),
        "PageNumber": 1,
        "Publisher": "false",
        "ReportingType": -1,
        "Symbol": symbol,
        "TracingNo": -1,
        "search": "true",
    }

    if from_date:
        params["FromDate"] = from_date

    if to_date:
        params["ToDate"] = to_date

    api_url="https://search.codal.ir/api/search/v2/q?"+ urlencode(params)

    # ---------------------------------------
    # create a session a find first page url
    # --------------------------------------

    session = create_session()
    response=session.get(url=api_url,timeout=30)
    response.raise_for_status()

    data=response.json()
    total_pages=data["Page"]
    page_urls=[]

    # -------------------
    # first pages
    # ------------------

    for letter in data["Letters"]:
        item = {
            "title": letter["Title"],
            "url": "https://www.codal.ir" + letter["Url"],
            "pdf_url":"https://www.codal.ir"+"/" + letter["PdfUrl"],
            "excel_url": letter["ExcelUrl"]
        }

        if letter["HasAttachment"]:
            item["attachment"] = "https://www.codal.ir" + letter["AttachmentUrl"]
        page_urls.append(item)

    # -------------------------------------------
    # other page because codal return first page
    # -------------------------------------------
    if total_pages>1:
        for page in range(2,total_pages+1):

            page_url = api_url.replace( "PageNumber=1", f"PageNumber={page}" )
            response = session.get(
            page_url,
            timeout=30
            )
            response.raise_for_status()

            data = response.json()

            for letter in data["Letters"]:
                item = {
                    "title": letter["Title"],
                    "url": "https://www.codal.ir" + letter["Url"],
                    "pdf_url":"https://www.codal.ir"+"/" + letter["PdfUrl"],
                    "excel_url": letter["ExcelUrl"]
                }

                if letter["HasAttachment"]:
                    item["attachment"] = "https://www.codal.ir" + letter["AttachmentUrl"]

                page_urls.append(item)

    # -----------------------
    # find the sheet id
    # ------------------------
    option = {
    "صورت سود و زیان تلفیقی": 13,
    "صورت سود و زیان جامع تلفیقی": 1097,
    "صورت وضعیت مالی تلفیقی": 14,
    "صورت تغییرات در حقوق مالکانه تلفیقی": 1099,
    "صورت جریان های نقدی تلفیقی": 15,

    "صورت سود و زیان": 1,
    "صورت سود و زیان جامع": 1058,
    "صورت وضعیت مالی": 0,
    "صورت تغییرات در حقوق مالکانه": 1060,
    "صورت جریان های نقدی": 9,
    }
    statement_name=statement_name.strip()
    sheet_id=option.get(statement_name)
    if sheet_id is None:
        raise ValueError(
        f"Sheet ID برای '{statement_name}' پیدا نشد."
        )

    # ---------------------------------
    # workers
    # --------------------------------

    all_datasources = []
    failed_reports = []
    no_datasource_reports = []
    decode_errors = []

    print(f"\nتعداد گزارش‌ها برای دریافت: {len(page_urls)}")

    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = []

        for item in page_urls:
            future = executor.submit(get_datasource,item,sheet_id)
            attachment=item.get("attachment")
            futures.append(future)
        for future in as_completed(futures):
            result= future.result()
            status = result["status"]

            if status == "success":
                all_datasources.append(result)
                print(f"✅ {result['title']}")

            elif status == "no_datasource":
                no_datasource_reports.append(result)
                print(f"⚠️ {result['title']} → datasource ندارد")

            elif status == "failed":
                failed_reports.append(result)
                print(f"❌ {result['title']}")
                print(f"{result['error']}")

            elif status == "decode_error":
                decode_errors.append(result)
                print(f"❌ Decode Error:{result['title']}" )
                print(f"{result['error']}")

    # --------------------------------------------------
    # results
    # --------------------------------------------------

    print("\n" + "-" * 50)

    print(" Total Pages:",total_pages)

    print("Total Report:",len(page_urls))

    print("Suceessfull:",len(all_datasources))

    print("Fail :",len(failed_reports))

    print(" datasource:",len(no_datasource_reports))

    print(" Decode:",len(decode_errors))

    print("-" * 50)

    return all_datasources





def report_to_workbook(all_datasource):
    """
    Convert extracted Codal reports into an Excel workbook.

    Args:
        all_datasource (list[dict]): Financial reports with datasource,
            metadata, and report links.

    Returns:
        openpyxl.Workbook: Excel workbook containing financial statements
            and report information.
    """

    #------------------------
    #sort the dataresource
    #------------------------

    all_datasource = sorted(
        all_datasource,
        key=lambda x: x["datasource"]["periodEndToDate"],
        reverse=True
    )
    wb = Workbook()
    ws = wb.active

    current_col = 2

    #Because codal report change before 1398
    old_description_created = False


    report_columns = []

    for i, report in enumerate(all_datasource):

        period = report["datasource"]["periodEndToDate"]
        year = int(period.split("/")[0])

        cells = (
            report["datasource"]
            ["sheets"][0]
            ["tables"][0]
            ["cells"]
        )

        # -----------------------------------------------
        # create first report
        # -----------------------------------------------
        if i == 0:

            for cell in cells:

                if cell["address"].startswith("A"):
                    ws[cell["address"]] = cell["value"]

            # the value
            report_col = current_col

            for cell in cells:

                if cell["address"].startswith("B"):

                    row = int(cell["address"][1:])

                    ws.cell(
                        row=row,
                        column=report_col
                    ).value = cell["value"]

            report_columns.append(report_col)

            current_col += 1


        # ------------------------------------------------
        # create others report greather than 1398
        # ----------------------------------------------
        elif year >= 1398:

            report_col = current_col

            for cell in cells:

                if cell["address"].startswith("B"):

                    row = int(cell["address"][1:])

                    ws.cell(
                        row=row,
                        column=report_col
                    ).value = cell["value"]

            report_columns.append(report_col)

            current_col += 1


        # ------------------------------------------
        # create others report less than 1398
        # ------------------------------------------
        else:
            # create description
            if not old_description_created:


                description_col = current_col

                for cell in cells:

                    if cell["address"].startswith("A"):

                        row = int(cell["address"][1:])

                        ws.cell(
                            row=row,
                            column=description_col
                        ).value = cell["value"]

                old_description_created = True

                current_col += 1


            # create other report less 1398
            report_col = current_col

            for cell in cells:

                if cell["address"].startswith("B"):

                    row = int(cell["address"][1:])

                    ws.cell(
                        row=row,
                        column=report_col
                    ).value = cell["value"]

            report_columns.append(report_col)

            current_col += 1


    # -------------------------------------------
    # last report about link of statement
    # -------------------------------------------

    title_col = current_col
    link_col = current_col + 1
    attachment_col = current_col + 2

    ws.cell(row=1, column=title_col).value = "عنوان گزارش"
    ws.cell(row=1, column=link_col).value = "لینک گزارش"
    ws.cell(row=1, column=attachment_col).value = "یاداشت های توضیحی"


    # -------------------------------------------
    # create links
    # -------------------------------------------

    for i, report in enumerate(all_datasource):

       #title report
        ws.cell(
            row=i + 2,
            column=title_col
        ).value = report["title"]

        # url report
        ws.cell(
            row=i + 2,
            column=link_col
        ).value = report["url"]

        # attachment url
        ws.cell(
            row=i + 2,
            column=attachment_col
        ).value = report["attachment"]


    return wb



def report_to_dataframe(all_datasource):
    """
        Convert financial reports into a pandas DataFrame.

    Args:
        all_datasource (list[dict]): Extracted financial reports.

    Returns:
            pandas.DataFrame: Financial report data.
    """

    wb=report_to_workbook(all_datasource)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    df=pd.read_excel(buffer)
    return df







def report_to_excel(all_datasource,path="E:/Report.xlsx"):
    """
    Export financial reports to a formatted Excel file.

    Args:
        all_datasource (list[dict]): Extracted financial reports.
        path (str): Output Excel file path.

    Returns:
        None: Saves the workbook to the specified path.
    """

    wb=report_to_workbook(all_datasource)
    ws=wb.active

    for row in ws.iter_rows():

        for cell in row:

            if isinstance(cell.value, str):

                try:

                    cell.value = int(
                        cell.value.replace(",", "")
                    )

                except:

                    pass

            if isinstance(
                cell.value,
                (int, float)
            ):

                cell.number_format = (
                    '#,##0;[Red]-#,##0'
                )

    font = Font(
        name="B Nazanin",
        size=10
    )

    for row in ws.iter_rows():

        for cell in row:

            cell.font = font

    ws.freeze_panes = "B3"

    blue_fill = PatternFill(
        fill_type="solid",
        start_color="00B0F0",
        end_color="00B0F0"
    )

    white_font = Font(
        name="B Nazanin",
        size=12,
        color="FFFFFF",
        bold=True
    )

    for row in (1, 2):

        for cell in ws[row]:

            cell.fill = blue_fill

            cell.font = white_font

    ws.title = "گزارش"

    for column in ws.columns:

        max_length = 0

        letter = get_column_letter(
            column[0].column
        )

        for cell in column:

            if cell.value is not None:

                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        ws.column_dimensions[
            letter
        ].width = max(
            10,
            min(
                max_length + 3,
                30
            )
        )

    start_link_col = ws.max_column - 1

    for row in range(
        2,
        ws.max_row + 1
    ):

        for col in range(
            start_link_col,
            ws.max_column + 1
        ):

            cell = ws.cell(
                row=row,
                column=col
            )

            if cell.value:

                url = str(
                    cell.value
                ).strip()

                if (
                    url.startswith("https://")
                    or
                    url.startswith("http://")
                ):

                    # متن نمایشی
                    cell.value = (
                        "https://www.codal.ir/Reports"
                    )

                    # لینک واقعی
                    cell.hyperlink = url

                    cell.style = "Hyperlink"

    ws_note = wb.create_sheet(
        "یاداشت های توضیحی"
    )

    # سه ستون آخر
    start_col = ws.max_column - 2

    for row in range(
        1,
        ws.max_row + 1
    ):

        for i in range(3):

            src = ws.cell(
                row=row,
                column=start_col + i
            )

            dst = ws_note.cell(
                row=row,
                column=i + 1
            )

            # مقدار
            dst.value = src.value

            # استایل
            if src.has_style:

                dst._style = copy(
                    src._style
                )

            # Hyperlink
            if src.hyperlink:

                dst.hyperlink = (
                    src.hyperlink.target
                )

                dst.style = "Hyperlink"

    for i in range(3):

        source_col = get_column_letter(
            start_col + i
        )

        target_col = get_column_letter(
            i + 1
        )

        ws_note.column_dimensions[
            target_col
        ].width = (
            ws.column_dimensions[
                source_col
            ].width
        )

    ws.delete_cols(
        start_col,
        3
    )

    wb.save(path)










# --------------------------------------------------------
# Settings
# --------------------------------------------------------

BASE_URL = "https://www.codal.ir/Reports/"

headers = {
    "User-Agent": (
        "Mozilla/5.0 (Linux; Android 15; Pixel 9) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/149.0.0.0 Mobile Safari/53.36"
    )
}


# --------------------------------------------------------
# Session
# --------------------------------------------------------

def create_session():
    """
    Create an HTTP session with automatic retries and default headers.

    Returns:
        requests.Session: Configured HTTP session.
    """

    session = requests.Session()

    retry = Retry(
        total=5,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET"],
        respect_retry_after_header=True
    )

    adapter = HTTPAdapter(max_retries=retry)

    session.mount("https://", adapter)
    session.headers.update(headers)

    return session


# --------------------------------------------------------
# Get attachment links
# --------------------------------------------------------

def get_attachment_links(report):
    """
    Extract attachment links from a Codal financial report.

    Args:
        report (dict): Financial report data and attachment URL.

    Returns:
        dict: Processing status, attachment links, and errors.
    """

    session = create_session()

    url = report.get("attachment")

    if not url:
        return {
            "status": "failed",
            "title": report.get("title"),
            "date": None,
            "links": [],
            "errors": [{
                "type": "attachment",
                "error": "Attachment URL وجود ندارد"
            }]
        }


    # Date
    try:
        date = report["datasource"]["periodEndToDate"]
        date = date.replace("/", "-") if date else None
    except (KeyError, TypeError):
        date = None


    # Request
    try:
        response = session.get(url, timeout=30)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        return {
            "status": "failed",
            "title": report.get("title"),
            "date": date,
            "links": [],
            "errors": [{
                "type": "request",
                "error": str(e)
            }]
        }


    # Parse
    try:
        soup = BeautifulSoup(response.text, "lxml")
        rows = soup.select("#dgAttachmentList tr.GridItem")


    except Exception as e:
        return {
            "status": "failed",
            "title": report.get("title"),
            "date": date,
            "links": [],
            "errors": [{
                "type": "parse",
                "error": str(e)
            }]
        }

    if not rows:
        return {
            "status": "failed",
            "title": report.get("title"),
            "date": date,
            "links": [],
            "errors": [{
                "type": "attachment",
                "error": "هیچ attachment ای پیدا نشد"
            }]
        }


    # Extract links
    links = []
    errors = []

    for row in rows:

        onclick = row.get("onclick", "")

        match = re.search(r"window\.open\(['\"]([^'\"]+)",onclick)

        if not match:
            errors.append({
                "type": "link",
                "error": "URL پیدا نشد"
            })
            continue

        full_url = urljoin(BASE_URL,match.group(1))

        description = row.select_one("td:nth-of-type(2)")

        description = (
            description.get_text(strip=True)
            if description else None)

        links.append({
            "date": date,
            "title": description,
            "page_url": url,
            "pdf": report.get("pdf"),
            "excel": report.get("excel"),
            "attachment_url": full_url
        })

    # Status
    if links and errors:
        status = "partial"
    elif links:
        status = "success"
    else:
        status = "failed"

    return {
        "status": status,
        "title": report.get("title"),
        "date": date,
        "links": links,
        "errors": errors
    }


# ---------------------------------------------------------------
# Multithreading
# --------------------------------------------------------------
def get_link(all_datasource):
        successful_reports = []
        partial_reports = []
        failed_reports = []
        results = []

        total = len(all_datasource)

        print(f"شروع پردازش {total} گزارش...")
        print("-" * 60)


        with ThreadPoolExecutor(max_workers=10) as executor:

            futures = [
                executor.submit(get_attachment_links, report)
                for report in all_datasource
            ]

            for completed, future in enumerate(
                as_completed(futures), 1
            ):

                try:

                    result = future.result()
                    results.append(result)

                    status = result["status"]

                    if status == "success":

                        successful_reports.append(result)

                        print(
                            f"✅ [{completed}/{total}] "
                            f"{result['date']} | "
                            f"{result['title']} | "
                            f"{len(result['links'])} attachment"
                        )

                    elif status == "partial":

                        partial_reports.append(result)

                        print(
                            f"⚠️ [{completed}/{total}] "
                            f"PARTIAL | "
                            f"{result['title']} | "
                            f"{len(result['links'])} attachment"
                        )

                    else:

                        failed_reports.append(result)

                        print(
                            f"❌ [{completed}/{total}] "
                            f"FAILED | "
                            f"{result['title']}"
                        )

                        for error in result["errors"]:

                            print(
                                f"   ❌ {error['type']}: "
                                f"{error['error']}"
                            )

                except Exception as e:


                    print(
                        f"❌ [{completed}/{total}] "
                        f"Worker Error: {e}"
                    )


        # --------------------------------------------------------------
        # Final result
        # --------------------------------------------------------------

        total_links = sum(
            len(result["links"])
            for result in results
        )

        print("\n" + "-" * 60)
        print(f"Total Report: {total}")
        print(f"Total link: {total_links}")
        print(f"Sucessfull: {len(successful_reports)}")
        print(f"Partial: {len(partial_reports)}")
        print(f"Fail: {len(failed_reports)}")

        print("\n" + "-" * 60)


        return  results


#version 1


# -------------------------------------------------------------
# settings
# -------------------------------------------------------------

WORKERS = 10

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/151.0.0.0 Safari/537.36"
    )
}


# -------------------------------------------------------------
# create session
# -------------------------------------------------------------


def create_session():
    """
    Create an HTTP session with retry support and default headers.

    Returns:
        requests.Session: Configured HTTP session.
    """

    session = requests.Session()

    retry = Retry(
        total=5,
        connect=5,
        read=5,
        status=5,
        backoff_factor=1,
        status_forcelist=[429,500,502,503,504],
        allowed_methods=["GET"],
        respect_retry_after_header=True)

    adapter = HTTPAdapter(max_retries=retry)

    session.mount("http://",adapter)

    session.mount("https://",adapter)

    session.headers.update(HEADERS)

    return session


# -------------------------------------------------------------
# clean file name should delete invalid character
# -------------------------------------------------------------

def clean_filename(name):
    """
    Sanitize a filename by removing invalid characters.

    Args:
        name (str): Original filename.

    Returns:
        str: Sanitized filename.
    """

    if not name:
        name = "attachment"

    # invalid character
    name = re.sub(r'[<>:"/\\|?*]',"_",name)

    # delete the space
    name = re.sub(r"\s+"," ",name).strip()

    # delete space and point the end of file
    name = name.rstrip(". ")

    return name


# -------------------------------------------------------------
# find the attechment type from server response
# -------------------------------------------------------------

def get_extension(response):

    """
    Determine the file extension from the response content type.

    Args:
        response (requests.Response): HTTP response object.

    Returns:
        str: Detected file extension.
    """

    content_type = response.headers.get("Content-Type","").lower()


    # pdf
    if "application/pdf" in content_type:

        return ".pdf"

    # excel
    if ("spreadsheetml" in content_type  or "application/vnd.ms-excel" in content_type):

        if "spreadsheetml" in content_type:
            return ".xlsx"

        return ".xls"


    # Word
    if "wordprocessingml" in content_type:

        return ".docx"

    if "application/msword" in content_type:

        return ".doc"


    # ZIP
    if "application/zip" in content_type:

        return ".zip"


    # RAR
    if ("application/x-rar" in content_type or "application/vnd.rar" in content_type):

        return ".rar"


    # Text
    if "text/plain" in content_type:

        return ".txt"


    # CSV
    if "text/csv" in content_type:

        return ".csv"


    # Unknown
    return ""


# -------------------------------------------------------------
# unique file name
# -------------------------------------------------------------

def get_unique_path(path):
    """
    Generate a unique file path if the given path already exists.

    Args:
        path (Path): Target file path.

    Returns:
        Path: Unique file path.
    """

    if not path.exists():

        return path

    counter = 1

    while True:

        new_path = (path.parent /f"{path.stem} ({counter}){path.suffix}")

        if not new_path.exists():

            return new_path

        counter += 1


# -------------------------------------------------------------
# DOWNLOAD ONE FILE
# -------------------------------------------------------------

def download_file(item):
    """
    Download a single file with retry and duplicate handling.

    Args:
        item (dict): Download information including URL and file path.

    Returns:
        dict: Download status, file information, and error details.
    """

    session = create_session()

    url = item["url"]

    target_path = Path(item["path"])

    file_type = item["type"]

    try:

        # -------------------------------------------------------------
        # if file already exits and is not empty skip
        # -------------------------------------------------------------

        if (target_path.exists() and target_path.stat().st_size > 0):

            return {
                "status": "skipped",
                "type": file_type,
                "url": url,
                "file": str(target_path),
                "message": "Already exists"
            }

        # -------------------------------------------------------------
        # request to server
        #-------------------------------------------------------------

        response = session.get(url,timeout=60)

        response.raise_for_status()

        # -------------------------------------------------------------
        # if have attachment file we can use "get_extention"  to find format of file
        # We can avoid duplicate names with "get_unique_path"
        # -------------------------------------------------------------

        if file_type == "attachment":

            extension = get_extension(response)

            # if find the format ass to path
            if extension:

                target_path = target_path.with_suffix(extension)

            # if file already exit
            target_path = get_unique_path(target_path)


        # create directory
        target_path.parent.mkdir(parents=True,exist_ok=True)

        # -------------------------------------------------------------
        # save a file in the form of chunk For optimal memory usage
        # chunk size is "1024*64 = 65536 ~ 66kb" if chunk is not empty write to file
        # -------------------------------------------------------------

        with open(target_path,"wb") as f:

            for chunk in response.iter_content(chunk_size=1024 * 64):

                if chunk:
                    f.write(chunk)

        # -------------------------------------------------------------
        # check the file status
        # -------------------------------------------------------------

        if (not target_path.exists()or target_path.stat().st_size == 0):

            return {
                "status": "failed",
                "type": file_type,
                "url": url,
                "file": str(target_path),
                "message": "Downloaded file is empty"
            }

        return {
            "status": "success",
            "type": file_type,
            "url": url,
            "file": str(target_path),
            "message": "Downloaded successfully"
        }

    except Exception as e:

        return {
            "status": "failed",
            "type": file_type,
            "url": url,
            "file": str(target_path),
            "message": str(e)
        }


# --------------------------------------------------------
# PREPARE DOWNLOADS
# --------------------------------------------------------

def prepare_downloads(all_datasource, base_folder):
    """
    Prepare a unique list of files to download.

    Args:
        results (list[dict]): Reports containing downloadable links.
        base_folder (str or Path): Base directory for downloaded files.

    Returns:
        list[dict]: Prepared download tasks.
    """
    results=get_link(all_datasource)
    base_folder = Path(base_folder)

    downloads = []


    # Preventing Duplicate URLs
    seen_urls = set()

    # Preventing Duplicate attachment name

    used_attachment_names = {}
    # --------------------------------------------------------
    # loop reports create a dowload list
    # --------------------------------------------------------

    for report in results:

        links = report.get("links",[])


        # LOOP LINKS
        for link in links:

            date = link.get("date","unknown_date")

            date = str(date)


            # create date folder
            date_folder = (base_folder /date)

            date_folder.mkdir(parents=True,exist_ok=True)

            # --------------------------------------------------------
            # PDF
            # --------------------------------------------------------

            pdf_url = link.get("pdf")

            if pdf_url:


                if pdf_url not in seen_urls:

                    seen_urls.add(pdf_url)

                    downloads.append({

                        "url": pdf_url,

                        "path": (
                            date_folder /
                            "صورت مالی.pdf"
                        ),

                        "type": "pdf",

                        "date": date,

                        "title": "صورت مالی"
                    })

            # --------------------------------------------------------
            # EXCEL
            # --------------------------------------------------------

            excel_url = link.get("excel")

            if excel_url:

                if excel_url not in seen_urls:

                    seen_urls.add(excel_url)

                    downloads.append({

                        "url": excel_url,

                        "path": (
                            date_folder /
                            "صورت مالی.xls"
                        ),

                        "type": "excel",

                        "date": date,

                        "title": "صورت مالی"
                    })

            # --------------------------------------------------------
            # ATTACHMENT
            # --------------------------------------------------------

            attachment_url = link.get("attachment_url")

            if attachment_url:


                if attachment_url in seen_urls:

                    continue

                seen_urls.add(attachment_url)

                # -------------------------------
                # title
                # -------------------------------

                title = clean_filename(link.get("title",""))

                if not title:

                    title = "attachment"

                # -------------------------------
                # dupilcate name
                # -------------------------------

                key = (date,title)

                if key not in used_attachment_names:

                    used_attachment_names[key] = 0

                    filename = title

                else:

                    used_attachment_names[key] += 1

                    number = (used_attachment_names[key])

                    filename = (f"{title} ({number})")

                downloads.append({

                    "url": attachment_url,

                    "path": (
                        date_folder /
                        filename
                    ),

                    "type": "attachment",

                    "date": date,

                    "title": title
                })

    return downloads


# --------------------------------------------------------------
# DOWNLOAD REPORTS
# --------------------------------------------------------------

def download_report(all_datasource, base_folder="E:/Codal_files"):
    """
    Download all report files concurrently.

    Args:
        results (list[dict]): Reports containing downloadable links.
        base_folder (str or Path): Base directory for downloaded files.

    Returns:
        dict: Download summary with successful, skipped, and failed files.
    """


    # -------------------------------
    # create dowload list
    # -------------------------------

    downloads = prepare_downloads(all_datasource,base_folder)





    # -------------------------------
    # THREAD POOL
    # -------------------------------

    successful = []
    skipped = []
    failed = []



    with ThreadPoolExecutor(max_workers=WORKERS) as executor:

        futures = {

            executor.submit(
                download_file,
                item
            ): item

            for item in downloads

        }



        for future in as_completed(futures):

            item = futures[future]

            try:

                result = future.result()

            except Exception as e:

                result = {

                    "status": "failed",

                    "type": item["type"],

                    "url": item["url"],

                    "file": str(
                        item["path"]
                    ),

                    "message": str(e)
                }


            # sucess
            if result["status"] == "success":

                successful.append(result)

                print(f"   {result['file']} ✅")

            # skipedd

            elif result["status"] == "skipped":

                skipped.append(result)

                print(f"  {result['file']} ⏭️")

            # failed

            else:

                failed.append(result)

                print(f" {result['file']} ❌")

                print(f"{result['message']}")

    # -----------------------------------------------------------
    # final report
    # -----------------------------------------------------------

    print()
    print("-" * 50)
    print("FINAL DOWNLOAD REPORT")
    print("-" * 50)

    print(f"Total:       {len(downloads)}")

    print(f"Successful:  {len(successful)}")

    print(f"Skipped:     {len(skipped)}")

    print(f"Failed:      {len(failed)}")

    print("-" * 50)


    # ------------------------------------------------------
    # failed file
    # ------------------------------------------------------

    if failed:

        print()
        print("-" * 70)
        print("FAILED DOWNLOADS")
        print("-" * 70)

        for item in failed:

            print()
            print(f"Type:    {item['type']}")

            print(f"File:    {item['file']}")

            print(f"URL:     {item['url']}")

            print(f"Error:   {item['message']}")

            print("-" * 70)

    # result
    return {

        "total": len(downloads),

        "successful": successful,

        "skipped": skipped,

        "failed": failed
    }

